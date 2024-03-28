import pandas as pd
import os

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Alignment,Font,PatternFill
from openpyxl.utils import get_column_letter

BILLS_FOLDER_PATH = os.path.join("bill docs")

def wage_calc_preprocessing(bill_pay_days,wages):
    wages["Days Present"] = [bill_pay_days["SKILLED"][0],bill_pay_days["SEMI-SKILLED"][0],bill_pay_days["UNSKILLED"][0]]
    wages["Days NH"] = [bill_pay_days["SKILLED"][1],bill_pay_days["SEMI-SKILLED"][1],bill_pay_days["UNSKILLED"][1]]
    wages = wages.reindex(wages.index[::-1]).reset_index(drop=True)
    wages["SL NO"] = [int(val)+1  for val in wages.index]
    wages = pd.concat([wages.iloc[:,-1:],wages.iloc[:,:-1]],axis=1)
    wages["mandays_pf"] = wages["Days Present"]
    wages["total_gross_amount"] = wages["Wage per day"] * wages["Days Present"]
    wages["total_nh_amount"] = wages["Wage per day"] * wages["Days NH"] 
    return wages


def generate_wage_calc_sheet(month,year,wage_calc_df,bill_pay_days,service_charges,ld,penalty,taxes,other_recovery,contract_no,vendor_name):
    
    template_folder = os.path.join(BILLS_FOLDER_PATH, "bill_templates")
    template_path = os.path.join(template_folder, "attendance.xlsx")

    month_year = pd.Period(f"{year}-{month}").start_time.date().strftime("%b-%Y")

    contract_line = f"Contract Reference: {contract_no}"
    vendor_line = f"Service Provider: {vendor_name}"
    month_line = f"Month: {month_year}"

    pf,edli,esi = bill_pay_days["EMPL_PF"],bill_pay_days["EMPL_EDLI"],bill_pay_days["EMPL_ESI"]
    nh_total = wage_calc_df["total_nh_amount"].sum()
    service_charges = bill_pay_days["SKILLED"][0] * service_charges[0] + bill_pay_days["SEMI-SKILLED"][0] * service_charges[1] +bill_pay_days["UNSKILLED"][0] * service_charges[2]
    
    workbook = load_workbook(template_path)
    sheet = workbook['Wage_Calculation']  # Update with your sheet's name
    sheet.column_dimensions["B"].width = 35
    
    cell = sheet.cell(row=4,column=1)
    cell.value = contract_line

    cell = sheet.cell(row=5,column=1)
    cell.value = vendor_line

    cell = sheet.cell(row=6,column=1)
    cell.value = month_line
    
    start_row = 8
    start_column = 3

    for index, row in wage_calc_df[["Wage per day","Days Present","mandays_pf","total_gross_amount"]].iterrows():
        for col_num, value in enumerate(row, start=start_column): 

            cell = sheet.cell(row=start_row + index, column=col_num)
            cell.value = value
            
    start_row = 13
    start_column = 3

    for index, row in wage_calc_df[["Wage per day","Days NH","Days NH","total_nh_amount"]].iterrows():
        for col_num, value in enumerate(row, start=start_column): 

            cell = sheet.cell(row=start_row + index, column=col_num)
            cell.value = value
            
    cell = sheet.cell(row=16,column=6)
    cell.value = pf

    cell = sheet.cell(row=17,column=6)
    cell.value = esi

    cell = sheet.cell(row=18,column=6)
    cell.value = edli

    cell = sheet.cell(row=20,column=6)
    cell.value = service_charges
    
    sub_total = (wage_calc_df.sum()["total_gross_amount"] + pf + esi + edli + service_charges + nh_total) 
    gst = sub_total * 0.18
    final_total = sub_total + gst
    
    cell = sheet.cell(row=21,column=6)
    cell.value = gst

    cell = sheet.cell(row=22,column=6)
    cell.value = final_total

    cell = sheet.cell(row=23,column=6)
    cell.value = final_total

    cell = sheet.cell(row=23,column=6)
    cell.value = ld

    cell = sheet.cell(row=24,column=6)
    cell.value = penalty

    cell = sheet.cell(row=25,column=6)
    cell.value = taxes

    cell = sheet.cell(row=26,column=6)
    cell.value = other_recovery
    
    total_recovery = ld + penalty + taxes + other_recovery
    cell = sheet.cell(row=27,column=6)
    cell.value = total_recovery
    
    net_payable = final_total - total_recovery

    cell = sheet.cell(row=28,column=6)
    cell.value = net_payable
    
    return sheet
    

def create_wage_calc_sheet(month,year,contract_no,sheet):

    attendance_folder = os.path.join(BILLS_FOLDER_PATH, contract_no, "attendance")
    target_path = os.path.join(attendance_folder, f"attendance_format_{month}_{year}.xlsx")
    attendance_workbook = load_workbook(target_path)

    try:
        if attendance_workbook["Wage_Calculation"]:
            attendance_workbook.remove(attendance_workbook["Wage_Calculation"])
            target_sheet = attendance_workbook.create_sheet(title = "Wage_Calculation")
    except Exception:
        target_sheet = attendance_workbook.create_sheet(title = "Wage_Calculation")

    for row in sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    for row in range(1,sheet.max_row+1):
        for col in range(1,sheet.max_column+1):
            source_cell = sheet.cell(row=row,column=col)
            target_cell = target_sheet.cell(row=row,column=col)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)

    for col_idx,column in enumerate(sheet.columns,start=1):
        target_sheet.column_dimensions[get_column_letter(col_idx)].width = sheet.column_dimensions[get_column_letter(col_idx)].width
    
    for row in sheet.iter_rows():
        for cell in row:
            target_sheet.row_dimensions[cell.row].height = sheet.row_dimensions[cell.row].height
            
    for merged_cell_range in sheet.merged_cells.ranges:
        target_sheet.merge_cells(merged_cell_range.coord)
        for merged_cell in merged_cell_range.cells:
            target_cell = target_sheet.cell(row=merged_cell[0],column=merged_cell[1])
            if merged_cell[0]  in [1,2]:
                # target_cell.alignment = Alignment(horizontal="left",vertical="center")
                pass
            # elif merged_cell[0]  in [4,5]:
            #     target_cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)               
            else:
                target_cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)

    attendance_workbook.save(target_path)
    return target_path