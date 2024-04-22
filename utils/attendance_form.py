import pandas as pd
import os

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Alignment,Font,PatternFill


BILLS_FOLDER_PATH = os.path.join("bill docs")
ATTENDANCE_TEMPLATE_PATH = "bill_templates/attendance.xlsx"



def custom_sort(skill_category):
    if skill_category == "SKILLED":
        return 1
    elif skill_category == "SEMI-SKILLED":
        return 2
    else:
        return 3

def map_values(value):
    if value == 'PP':
        return 1
    elif value == 'AA':
        return 0
    elif value in ["AP","PA"] :  # Assuming PA or AP
        return 0.5  
    else:
        return value

def read_punching_data(nh_days):
    attendance_df = pd.read_csv("utils/updated_punch_data.csv")
    # print(str(attendance_df))
    attendance_df = attendance_df.map(map_values)
    # print(str(attendance_df))
    attendance_df["NH Days"] = nh_days

    df_columns = attendance_df.columns.tolist()

    sunday_columns = []
    for col in df_columns:
        try:
            if datetime.strptime(col, "%d/%m").weekday() == 5:  # 6 is Sunday
                sunday_columns.append(col)

        except ValueError:
            pass
    
    attendance_df.loc[:,sunday_columns] = "S"
    return attendance_df,sunday_columns


def attendance_processing(month, year, contract_no, emp_categories):
    # Read csv file
    attendance_folder = os.path.join(BILLS_FOLDER_PATH, contract_no, "attendance")
    csv_file_path = os.path.join(attendance_folder, f"{month}-{year}.csv")
    attendance_df = pd.read_csv(csv_file_path).set_index("Unnamed: 0")
    employees = attendance_df.columns[:-2]  # Extract employees
    attendance_df.index = range(1, len(attendance_df) + 1)  # set index from 1

    # add last row with index as "TOTAL PAY DAYS" and sum of days present
    last_row = []
    for col in attendance_df.columns:
        if col in employees:
            total = attendance_df.agg({col: "sum"}).values[0]
            last_row.append(total)
        elif col == "date_format":
            last_row.append("DP")
        else:
            last_row.append("")
    attendance_df.loc["TOTAL PAY DAYS"] = last_row

    # change sundays with 0 values to S
    for col in employees:
        attendance_df.loc[(attendance_df.sunday == 'sunday') & (attendance_df[col] == 0.0), col] = "S"

    # get sunday list for yellow color in excel sheet
    sunday_list = [val for val in attendance_df.sunday.values if val != ""]

    # drop sundat and date_format columns
    attendance_df.drop(["sunday", "date_format"], axis=1, inplace=True)

    # get data to required format
    result_df = attendance_df.T

    # fill balance days to 31 days with X
    result_days, result_total = result_df[result_df.columns[:-1]], result_df[result_df.columns[-1:]]

    days = result_days.shape[1]
    days_to_add = 31 - days

    for i in range(days_to_add):
        result_days[days + i + 1] = 'X'
        sunday_list.append("")

    # join result_days and result_total
    result_df = result_days.join(result_total).reset_index().rename(columns={'index': "Emp Name"})

    # create NH DAY and SL.No columns and reindex the columns
    result_df["NH DAY"] = 1.0
    result_df["CATEGORY"] = emp_categories
    result_df = pd.concat([result_df.iloc[:, -1:], result_df.iloc[:, :-1]], axis=1)
    result_df = result_df.sort_values(by="CATEGORY",key=lambda x:x.map(custom_sort))
    result_df["SL.No"] = range(1, len(result_df) + 1)
    result_df = pd.concat([result_df.iloc[:, -1:], result_df.iloc[:, :-1]], axis=1).reset_index(drop=True)
    print(str(result_df))

    return result_df, sunday_list


def generate_attendance_excel(result_df, month, year, work_description,contract_no,
                               dept_name, dept_intercom, vendor_name,
                               work_order_no, gem_contract_no, contract_st_dt,
                               sunday_list,total_pay_days):

    template_folder = os.path.join(BILLS_FOLDER_PATH, "bill_templates")
    attendance_folder = os.path.join(BILLS_FOLDER_PATH, contract_no, "attendance")
    template_path = os.path.join(template_folder, "attendance.xlsx")
    target_path = os.path.join(attendance_folder, f"attendance_format_{month}_{year}.xlsx")

    month_st_dt = pd.Period(f"{year}-{month}").start_time.date().strftime("%d-%m-%Y")
    month_end_dt = pd.Period(f"{year}-{month}").end_time.date().strftime("%d-%m-%Y")


    contract_end_dt = (pd.to_datetime(contract_st_dt) + pd.DateOffset(months=24,days=-1)).date().strftime("%d-%m-%Y")
    contract_st_dt = pd.to_datetime(contract_st_dt).date().strftime("%d-%m-%Y")

    workbook = load_workbook(template_path)
    max_name_length = max([len(name) for name in result_df["Emp Name"].values])

    sheet = workbook['ATTENDANCE']  # Update with your sheet's name
    sheet.column_dimensions["C"].width = (max_name_length + 2) * 1.2
    sheet.column_dimensions["B"].width = 15

    border_style = Border(left=Side(border_style="thin", color="FF000000"),
                          right=Side(border_style="thin", color="FF000000"),
                          top=Side(border_style="thin", color="FF000000"),
                          bottom=Side(border_style="thin", color="FF000000"))

    alignment_style = Alignment(horizontal='center', vertical="center")
    alignment_style_left = Alignment(horizontal='left', vertical="center")
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    cell = sheet.cell(row=2, column=25)
    cell.value = month_st_dt

    cell = sheet.cell(row=2, column=34)
    cell.value = month_end_dt

    cell = sheet.cell(row=3, column=4)
    cell.value = work_description

    cell = sheet.cell(row=4, column=4)
    cell.value = dept_name

    cell = sheet.cell(row=4, column=28)
    cell.value = dept_intercom

    cell = sheet.cell(row=5, column=4)
    cell.value = vendor_name

    cell = sheet.cell(row=6, column=4)
    cell.value = work_order_no

    cell = sheet.cell(row=6, column=27)
    cell.value = gem_contract_no

    cell = sheet.cell(row=7, column=25)
    cell.value = contract_st_dt

    cell = sheet.cell(row=7, column=34)
    cell.value = contract_end_dt

    start_row = 9
    start_column = 1  # Assuming you start from column A

    sheet.delete_rows(start_row, 38)

    for index, row in result_df.iterrows():
        for col_num, value in enumerate(row, start=start_column):
            cell = sheet.cell(row=start_row + index, column=col_num)
            cell.value = value
            cell.font = bold_font
            if col_num > 3 and col_num < 35 : 
                if sunday_list[col_num - 4] == "sunday":
                    cell.fill = yellow_fill

            if col_num in [2, 3]:
                cell.alignment = alignment_style_left
            else:
                cell.alignment = alignment_style

    sheet.cell(row=start_row + len(result_df), column=1).value = '*'
    sheet.cell(row=start_row + len(result_df),
               column=2).value = '''"1"- Means Present, "0"- Means Absent, "0.5"-Means Halfday Present'''
    sheet.merge_cells(start_row=start_row + len(result_df), start_column=2, end_row=start_row + len(result_df),
                      end_column=34)

    row_no = start_row + len(result_df) + 1
    sheet.cell(row=row_no, column=1).value = "TOTAL"
    sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=34)

    cell = sheet.cell(row=row_no, column=35)
    cell.value = total_pay_days
    cell.alignment = alignment_style
    cell.font = bold_font

    cell = sheet.cell(row=row_no, column=36)
    cell.value = 0
    cell.alignment = alignment_style
    cell.font = bold_font

    row_no = start_row + len(result_df) + 2
    sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no + 3, end_column=9)
    sheet.merge_cells(start_row=row_no, start_column=10, end_row=row_no + 3, end_column=25)
    sheet.merge_cells(start_row=row_no, start_column=26, end_row=row_no + 3, end_column=36)

    row_no = start_row + len(result_df) + 6
    sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=9)
    sheet.merge_cells(start_row=row_no, start_column=10, end_row=row_no, end_column=25)
    sheet.merge_cells(start_row=row_no, start_column=26, end_row=row_no, end_column=36)
    sheet.cell(row=row_no, column=1).value = "PREPARED BY"
    sheet.cell(row=row_no, column=10).value = "ACCEPTED BY CONTRACTOR"
    sheet.cell(row=row_no, column=26).value = "HEAD OF DEPT."

    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=36):
        for cell in row:
            cell.border = border_style

    sheets_to_delete = [sheet for sheet in workbook.sheetnames if sheet != "ATTENDANCE"]

    for sheet in sheets_to_delete:
        workbook.remove(workbook[sheet])
    print(target_path)
    workbook.save(filename=target_path)
    return target_path
