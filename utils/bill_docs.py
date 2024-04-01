import os
import pandas as pd

from openpyxl import load_workbook



BILLS_FOLDER_PATH = os.path.join("bill docs")



def get_latest_excel_file(contract_no):

    folder_path = os.path.join(BILLS_FOLDER_PATH,contract_no,"attendance")
    # Get a list of all files in the folder
    files = os.listdir(folder_path)

    # Filter only Excel files (assuming they have a .xlsx extension)
    excel_files = [file for file in files if file.endswith('.xlsx')]

    # If there are no Excel files, return None
    if not excel_files:
        return None

    # Get the full paths of all Excel files
    excel_files_paths = [os.path.join(folder_path, file) for file in excel_files]

    # Get the latest modified Excel file based on modification timestamp
    latest_file = max(excel_files_paths, key=os.path.getmtime)

    folder_path, file_name = os.path.split(latest_file)

    return file_name


def get_from_and_to_dates(contract_no,selected_month):
    month,year = selected_month.split("-")
    filename = f"attendance_format_{month}_{year}.xlsx"
    file_path = os.path.join(BILLS_FOLDER_PATH,contract_no,"attendance",filename)
    from_date,to_date = pd.read_excel(file_path,sheet_name="ATTENDANCE").iloc[0,[24,33]].values.tolist()
    return from_date,to_date


def get_checklist_form_data(contract):
    contract_end_date = (pd.to_datetime(contract.start_date) + pd.DateOffset(months=contract.duration_months,days=-1)).date().strftime("%d-%m-%Y")
    contract_start_date = pd.to_datetime(contract.start_date).date().strftime("%d-%m-%Y")

    file_path = os.path.join(BILLS_FOLDER_PATH,"bill_templates","attendance.xlsx")
    checklist_df = pd.read_excel(file_path,sheet_name="CHECKLIST",header=6,skipfooter=5)
    checklist_df["Remarks"] = checklist_df["Remarks"].str.replace(r'\s+', ' ').fillna("NA")
    checklist_df["Documents"] = checklist_df["Documents"].str.replace("\n","")
    checklist_df.loc[[3,4],"Remarks"] = contract_start_date,contract_end_date
    checklist_form_data = checklist_df[["Documents","Remarks"]].to_dict(orient='records')
    return checklist_form_data

def get_undertaking_form_data():
    undertaking_form_data = ["Vendor's original invoice amount (including taxes)","Is original invoice amount and IFS entered invoice amount equal","Retention Money","Keepback","Penalty",
        "Other Deductions, if any","Rents (Quarter/Electricity/Water):Rs.",]

    return undertaking_form_data


def generate_checklist_sheet(checklist_data,contract,values):
    template_path = os.path.join(BILLS_FOLDER_PATH,"bill_templates","attendance.xlsx")

    rar_no = checklist_data.get("rar_no")
    contract_no = checklist_data.get("contract_no")

    rar_folder_path = os.path.join(BILLS_FOLDER_PATH,contract_no,f"RAR_{rar_no}")
    rar_file_path = os.path.join(rar_folder_path,f"RAR_{rar_no}.xlsx")
    invoice_no = pd.read_excel(rar_file_path,sheet_name="RAR").iloc[10,0].split("-")[1].strip()

    workbook = load_workbook(template_path)
    sheet = workbook["CHECKLIST"]

    cell = sheet.cell(row=2,column=1)
    work_description = f"SUBJECT   : {contract.contract_description}"
    cell.value = work_description

    cell = sheet.cell(row=3,column=1)
    vendor_name_line = f"CONTRACTOR   :{contract.vendor_name}"
    cell.value = vendor_name_line

    cell = sheet.cell(row=4,column=1)
    contract_no_line = f"CONTRACT NO   :{contract.contract_no}({contract.gem_contract_no})"
    cell.value = contract_no_line

    cell = sheet.cell(row=5,column=1)
    rar_line = f'''RAR/FINAL BILL   :{checklist_data.get("rr_no")},Dt: {checklist_data.get("ge_date")};Invoice No- {invoice_no},Dt: {checklist_data.get("ge_date")} (RAR-{checklist_data.get("rar_no")})'''
    cell.value = rar_line


    for row_index, value in enumerate(values, start=1):
        cell = f'C{row_index+7}'  # Construct cell reference (e.g., A1, A2, etc.)
        sheet[cell] = value

    target_path = os.path.join(rar_folder_path,"checklist.xlsx")
    workbook.save(target_path)
    return 
