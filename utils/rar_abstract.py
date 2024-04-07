import pandas as pd
import os

from data_model_flask_alchemy import db
from sqlalchemy.exc import SQLAlchemyError

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Alignment,Font,PatternFill
from openpyxl.utils import get_column_letter

BILLS_FOLDER_PATH = os.path.join("bill docs")

def get_rar_quantities(session_rar_abstract_data,descriptions):

    month,year = session_rar_abstract_data.get("selected_month").split("-")
    contract_no = session_rar_abstract_data.get("contract_no")

    attendance_folder = os.path.join(BILLS_FOLDER_PATH, contract_no, "attendance")
    wage_calc_file_path = os.path.join(attendance_folder, f"attendance_format_{month}_{year}.xlsx")

    wage_calc_df = pd.read_excel(wage_calc_file_path,sheet_name="Wage_Calculation",header=6)
    mandays = wage_calc_df.loc[:2,"Mandays for PF"].values
    pf_edli = wage_calc_df.loc[[8,10],"Total Gross Salary Amount"].sum(axis=0)
    esi = wage_calc_df.loc[[9],"Total Gross Salary Amount"].sum(axis=0)
    nh_wages = 0
    leave_wages = 0

    present_rar_qty = [0]*len(descriptions)
    if len(descriptions) > 10:
        present_rar_qty[3:10] = [pf_edli,esi,nh_wages,leave_wages] + list(mandays)
        present_rar_qty[-3:] = list(mandays)


    return present_rar_qty




if __name__ == "__main__":
    get_rar_quantities(session_rar_abstract_data = {"selected_month":"2-2024","contract_no":"22SNCJO-373"},descriptions = [1,2])
