import pandas as pd
import os

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Alignment,Font,PatternFill
from openpyxl.utils import get_column_letter

BILLS_FOLDER_PATH = os.path.join("bill docs")


def group_processing(group,category,prof_tax):
    group.drop(["SL NO"],axis=1,inplace=True)
    group["gross_wage"] = group["TOTAL PAY DAYS"] * group["Wage per day"] 
    group["nh_wage"] = group["NH DAY"] * group["Wage per day"] 
    group["gross_wage_pf"] = group["gross_wage"].apply(lambda x:15000 if x > 15000 else x)
    group["emp_epf"] = (group["gross_wage_pf"] * 0.12).round(2)
    group["emp_esi"] = (group["gross_wage"] * 0.0075).round(2)
    group["prof_tax"] =  group["gross_wage_pf"].apply(lambda x: prof_tax if x == 15000 else 0)
    group["emp_epf_esi_total"] = group["emp_epf"] + group["emp_esi"]
    group["empl_epf"] = (group["gross_wage_pf"] * 0.125).round(2)
    group["empl_edli"] = (group["gross_wage_pf"] * 0.005).round(2)
    group["empl_epf_edli_total"] = group["empl_epf"] + group["empl_edli"]
    group["empl_esi"] = (group["gross_wage"] * 0.0325).round(2)
    group["empl_epf_edli_esi_total"] = group["empl_epf"] + group["empl_edli"] + group["empl_esi"]
    group["net_pay"] = (group["gross_wage"] + group["nh_wage"]) - group["emp_epf_esi_total"] - group["prof_tax"]

    last_row = group.sum(axis=0).values
    last_row[:2] = ["",f"{category} TOTAL"]
    group.loc["total"] = last_row
    return group


def pf_esi_preprocessing(attendance_path,wages:pd.DataFrame,prof_tax):
    usecols=["SL NO","CATEGORY OF SKILLNESS","NAME OF CONTRACT PERSONNEL","TOTAL PAY DAYS","NH DAY"]
    data = pd.read_excel(attendance_path,header=7,usecols=usecols,skipfooter=7)
    
    merged_data = data.merge(wages,how="left",left_on="CATEGORY OF SKILLNESS",right_on="CATEGORY").drop(["CATEGORY"],axis=1)
    groups = merged_data.groupby("CATEGORY OF SKILLNESS")
    
    processed_df = pd.DataFrame()
    
    bill_pay_days ={}

    for category in merged_data["CATEGORY OF SKILLNESS"].unique():

        group = groups.get_group(category)
        processed_group = group_processing(group,category,prof_tax)
        bill_pay_days[category] = processed_group.loc["total",["TOTAL PAY DAYS","NH DAY"]].values.tolist()
        processed_df = pd.concat([processed_df,processed_group],axis=0)


    final_row = processed_df.loc["total"].sum(axis=0).values
    final_row[:2] = ["","TOTAL"]
    processed_df.loc["final_total"] = final_row
    processed_df.loc[["final_total","total"],["Wage per day","nh_wage"]] = ""

    processed_df["SL NO"] = [int(val)+1 if val not in ["total","final_total"] else "" for val in processed_df.index]
    processed_df = pd.concat([processed_df.iloc[:,-1:],processed_df.iloc[:,:-1]],axis=1)

    empl_epf,empl_edli,esi_total = processed_df.loc["final_total",["empl_epf","empl_edli","empl_esi"]].values
    bill_pay_days["EMPL_PF"] = empl_epf
    bill_pay_days["EMPL_EDLI"] = empl_edli
    bill_pay_days["EMPL_ESI"] = esi_total
    for category in ["SKILLED","SEMI-SKILLED","UNSKILLED"]:
        if category not in bill_pay_days.keys():
            bill_pay_days.update({category:[0.0,0.0]})

    processed_df.reset_index(drop=True,inplace=True)
    processed_df.drop(["CATEGORY OF SKILLNESS"],axis=1,inplace=True)

    rows_yellow_fill = processed_df[processed_df["NAME OF CONTRACT PERSONNEL"].str.contains("TOTAL")].index.values
    
    return processed_df,bill_pay_days,rows_yellow_fill


def generate_pf_esi_sheet(processed_df,rows_yellow_fill,month,year,vendor_name):
    '''
    This function reads attendance file and return styled excel sheet with values.
    '''
    
    template_folder = os.path.join(BILLS_FOLDER_PATH, "bill_templates")
    template_path = os.path.join(template_folder, "attendance.xlsx")
    
    month_st_dt = pd.Period(f"{year}-{month}").start_time.date().strftime("%d-%m-%Y")
    month_end_dt = pd.Period(f"{year}-{month}").end_time.date().strftime("%d-%m-%Y")
    month_year = pd.Period(f"{year}-{month}").start_time.date().strftime("%b-%Y")
    month_line = f"Due Wage Month:{month_year} from {month_st_dt} to {month_end_dt}"

    vendor_name_line = f"Name of Contractor : {vendor_name}"
    
    workbook = load_workbook(template_path)
    max_name_length = max([len(name) for name in processed_df["NAME OF CONTRACT PERSONNEL"].values])

    sheet = workbook['PF_ESI']  # Update with your sheet's name
    sheet.column_dimensions["B"].width = (max_name_length+2)*1.2

    border_style = Border(left=Side(border_style="thin",color="FF000000"),
                              right=Side(border_style="thin",color="FF000000"),
                              top=Side(border_style="thin",color="FF000000"),
                              bottom=Side(border_style="thin",color="FF000000"))

    alignment_style = Alignment(horizontal='center',vertical="center")
    alignment_style_left = Alignment(horizontal='left',vertical="center")
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
    green_fill = PatternFill(start_color="00FF00",end_color="00FF00",fill_type="solid")

    cell = sheet.cell(row=2,column=1)
    cell.value = month_line

    cell = sheet.cell(row=3,column=1)
    cell.value = vendor_name_line

    start_row = 6
    start_column = 1  # Assuming you start from column A

    sheet.delete_rows(start_row,sheet.max_row)

    for index, row in processed_df.iterrows():

        for col_num, value in enumerate(row, start=start_column):   

            cell = sheet.cell(row=start_row + index, column=col_num)
            cell.value = value
            cell.font = bold_font

            if col_num != 1:
                cell.number_format = "0.00"

            if start_row + index - 6  in rows_yellow_fill:
                cell.fill = yellow_fill

            if col_num == 2:
                cell.alignment = alignment_style_left
            else:
                cell.alignment = alignment_style

    last_row = sheet[sheet.max_row]
    for cell in last_row:
        cell.fill = green_fill

    for row in sheet.iter_rows(min_row=start_row,min_col=1,max_row=5+len(processed_df),max_col=18):
            for cell in row:
                cell.border = border_style

    return sheet


def create_pf_esi_sheet(month,year,contract_no,sheet):
    
    attendance_folder = os.path.join(BILLS_FOLDER_PATH, contract_no, "attendance")
    
    target_path = os.path.join(attendance_folder, f"attendance_format_{month}_{year}.xlsx")
    attendance_workbook = load_workbook(target_path)

    try:
        if attendance_workbook["PF_ESI"]:
            attendance_workbook.remove(attendance_workbook["PF_ESI"])
            target_sheet = attendance_workbook.create_sheet(title = "PF_ESI")
    except Exception:
        target_sheet = attendance_workbook.create_sheet(title = "PF_ESI")

    for row in sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    for row in range(1,sheet.max_row+1):
        for col in range(1,sheet.max_column+1):
            source_cell = sheet.cell(row=row,column=col)
            target_cell = target_sheet.cell(row=row,column=col)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)

    for col_idx,column in enumerate(sheet.columns,start=1):
        target_sheet.column_dimensions[get_column_letter(col_idx)].width = sheet.column_dimensions[get_column_letter(col_idx)].width

    for row in sheet.iter_rows():
        for cell in row:
            target_sheet.row_dimensions[cell.row].height = sheet.row_dimensions[cell.row].height

    for merged_cell_range in sheet.merged_cells.ranges:
        target_sheet.merge_cells(merged_cell_range.coord)
        for merged_cell in merged_cell_range.cells:
            target_cell = target_sheet.cell(row=merged_cell[0],column=merged_cell[1])
            if merged_cell[0]  in [2,3]:
                target_cell.alignment = Alignment(horizontal="left",vertical="center")
            elif merged_cell[0]  in [4,5]:
                target_cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            else:
                target_cell.alignment = Alignment(horizontal="center",vertical="center")

    attendance_workbook.save(target_path)
    return target_path